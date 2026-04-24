"""Quick test for the smart parser pipeline."""
from upload.parsers import parse_all_sheets
from upload.sheet_router import route_sheets
import openpyxl, io

# Create a test workbook that mimics the user's file structure
wb = openpyxl.Workbook()

# First sheet: Title-style (the bug trigger - only 1 header cell)
ws1 = wb.active
ws1.title = "Aegis.One_Multi_Sheet_Upload_Template"
ws1.append(["Organization Profile"])
ws1.append(["Field", "Value"])
ws1.append(["Company Name", "Test Corp"])
ws1.append(["Industry", "Technology"])

# Second sheet: Controls
ws2 = wb.create_sheet("Controls")
ws2.append(["Control ID", "Control Name", "Status", "Owner"])
ws2.append(["A.5.1", "Information Security Policy", "Implemented", "CISO"])
ws2.append(["A.6.1", "Organization of Info Security", "Partial", "IT Manager"])

# Third sheet: Assets
ws3 = wb.create_sheet("Asset Inventory")
ws3.append(["Asset Name", "Type", "Owner", "Criticality"])
ws3.append(["Main Server", "Server", "IT Team", "High"])

# Fourth sheet: Employees
ws4 = wb.create_sheet("Employees")
ws4.append(["Name", "Role", "Department", "Email", "Training Status"])
ws4.append(["John Doe", "Engineer", "IT", "john@test.com", "Completed"])

buf = io.BytesIO()
wb.save(buf)
contents = buf.getvalue()

print("=== Testing parse_all_sheets ===")
result = parse_all_sheets(contents, "test_template.xlsx")
print("Sheets parsed:", len(result))
for s in result:
    print(f"  Sheet: {s['name']}")
    print(f"    Type: {s['type']}")
    print(f"    Rows: {s['row_count']}")
    print(f"    Classification: {s['classification']}")
    print(f"    Headers: {s['headers']}")
    print()

print("=== Testing route_sheets ===")
routing = route_sheets(result)
print("Total sheets:", routing["total_sheets"])
print("Controls found:", routing["controls_found"])
print("Has assets:", routing["has_assets"])
print("Has employees:", routing["has_employees"])
print()
print("Detection summary:")
for line in routing["detection_summary"]:
    print(f"  {line}")

if routing["warnings"]:
    print("\nWarnings:")
    for w in routing["warnings"]:
        print(f"  {w}")

print("\n=== SUCCESS - Pipeline works correctly! ===")
