"""
Export Service.

Generates Excel workbook and PDF report from the latest completed
assessment stored in data/assessments.json.
"""

import io
import json
import os
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
_ASSESSMENTS_FILE = os.path.join(_DATA_DIR, "assessments.json")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_assessments() -> list[dict]:
    """Load all assessment records from the JSON store."""
    if not os.path.exists(_ASSESSMENTS_FILE):
        return []
    with open(_ASSESSMENTS_FILE, "r", encoding="utf-8") as fh:
        return json.load(fh)


def get_latest_completed_assessment() -> dict | None:
    """
    Return the most recent assessment that contains a framework_assessment
    with populated data (compliance_score is not None).
    """
    records = _load_assessments()
    for rec in reversed(records):
        fa = rec.get("framework_assessment", {})
        if fa and fa.get("compliance_score") is not None:
            # Merge top-level metadata into the framework_assessment dict
            fa["_company_name"] = rec.get("assessment_name", "")
            fa["_file_name"] = rec.get("file_name", "")
            fa["_top_framework"] = rec.get("framework", "")
            fa["_top_created_at"] = rec.get("created_at", "")
            return fa
    return None


def _safe_filename(text: str) -> str:
    """Sanitise a string for use in a filename."""
    return re.sub(r"[^a-zA-Z0-9_\-]", "_", (text or "export").strip())[:60]


def _fmt_date(iso_str: str) -> str:
    """Format an ISO date string to YYYY-MM-DD, or return today."""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="111936", end_color="111936", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _add_header_row(ws, headers: list[str], row: int = 1):
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _write_rows(ws, rows: list[list], start_row: int = 2):
    """Write data rows with basic styling."""
    for r_idx, row_data in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    """Auto-adjust column widths."""
    for col in ws.columns:
        longest = min_width
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                longest = max(longest, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = longest + 2


def build_excel_workbook(fa: dict) -> tuple[io.BytesIO, str]:
    """
    Build an Excel workbook from a framework_assessment dict matching UI logic.
    Returns (BytesIO buffer, suggested filename).
    """
    import re
    wb = openpyxl.Workbook()
    company = _safe_filename(fa.get("assessment_name") or fa.get("_company_name") or "assessment")
    date_str = _fmt_date(fa.get("created_at") or fa.get("_top_created_at", ""))
    fw = str(fa.get("framework", ""))
    is_hipaa = "hipaa" in fw.lower()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    summary_data = [
        ["Field", "Value"],
        ["Assessment Name", fa.get("assessment_name", "")],
        ["Framework", fw],
        ["Date", date_str],
        ["Compliance Score", f"{fa.get('compliance_score', 0)}%"],
        ["Total Controls", fa.get("total_controls", 0)],
        ["Compliant", fa.get("compliant_controls", 0)],
        ["Partial", fa.get("partial_controls", 0)],
        ["Missing", fa.get("missing_controls", 0)],
        ["Evidence Backed", "Yes" if fa.get("evidence_backed") else "No"],
    ]
    _add_header_row(ws, summary_data[0])
    _write_rows(ws, summary_data[1:])
    _auto_width(ws)

    # --- Sheet 2: SoA ---
    soa_rows = []
    if is_hipaa and fa.get("soa", {}).get("entries"):
        for i, e in enumerate(fa.get("soa").get("entries", []), 1):
            control_no = e.get("control_no", "")
            m = re.search(r'[Aa](\d+\.\d+)', control_no)
            if m: control_no = f"A.{m.group(1)}"
            
            soa_rows.append([
                i,
                e.get("section", "—"),
                control_no,
                e.get("control_title", "—"),
                e.get("applicable", "Yes"),
                str(e.get("status", "missing")).upper(),
                e.get("implementation", e.get("remarks", "—")),
                e.get("reference", "—"),
            ])
    else:
        row_id = 1
        for section in fa.get("sections", []):
            for c in section.get("controls", []):
                status = str(c.get("status", "")).lower()
                
                implemented = "Control is not implemented; no supporting evidence found"
                if status == "compliant": implemented = "Control is implemented and supported by available evidence"
                elif status == "partial": implemented = "Control is partially implemented; improvements required"
                
                domain_hint = str(c.get("domain", "") + section.get("section_name", "") + section.get("section_key", "")).lower()
                rule_hint = str(c.get("rule_id", "")).lower()
                reference = "System Inference Engine"
                
                if status == "missing" and not c.get("has_evidence"):
                    reference = "System Inference Engine / No supporting evidence"
                elif "vendor" in domain_hint or "supplier" in domain_hint or "vendor" in rule_hint:
                    reference = "Vendor Records / Third-Party Data"
                elif "people" in domain_hint or "a6" in domain_hint or "iso-a6" in rule_hint:
                    reference = "HR / Employee Data; Training Matrix"
                elif "physical" in domain_hint or "a7" in domain_hint or "iso-a7" in rule_hint:
                    reference = "Asset Inventory / Facilities Data"
                elif "technological" in domain_hint or "a8" in domain_hint or "iso-a8" in rule_hint:
                    reference = "Network Rules / Configurations / Systems"
                elif "organizational" in domain_hint or "a5" in domain_hint or "iso-a5" in rule_hint:
                    reference = "Policies / Governance / Risk Register"
                
                control_no = str(c.get("rule_id", ""))
                m = re.search(r'[Aa](\d+\.\d+)', control_no)
                if m: control_no = f"A.{m.group(1)}"
                
                soa_rows.append([
                    row_id,
                    section.get("section_name", section.get("section_key", "—")),
                    control_no,
                    c.get("name", "—"),
                    "Yes",
                    status.upper(),
                    implemented,
                    reference
                ])
                row_id += 1

    if soa_rows:
        ws_soa = wb.create_sheet("SoA")
        headers = ["ID", "Section", "Control No", "Control Title", "Applicable", "Status", "Implementation", "Reference"]
        _add_header_row(ws_soa, headers)
        _write_rows(ws_soa, soa_rows)
        _auto_width(ws_soa)

    # --- Sheet 3: Compliance Matrix ---
    cm = fa.get("compliance_matrix")
    cm_rows = []
    cm_entries = []
    if isinstance(cm, dict):
        cm_entries = cm.get("entries", cm.get("controls", []))
    elif isinstance(cm, list):
        cm_entries = cm

    if cm_entries:
        ws_cm = wb.create_sheet("Compliance Matrix")
        first = cm_entries[0]
        if "Framework" in first and "Requirement" in first:
            headers = ["Framework", "Requirement", "Status", "Mapped Controls", "Gaps Identified", "Remediation Plan"]
            _add_header_row(ws_cm, headers)
            for row in cm_entries:
                cm_rows.append([
                    row.get("Framework", ""),
                    row.get("Requirement", ""),
                    row.get("Status", ""),
                    str(row.get("Mapped Controls", "")),
                    str(row.get("Gaps Identified", "")),
                    str(row.get("Remediation Plan", ""))
                ])
        else:
            headers = ["Control ID", "Control Name", "Status", "Evidence", "Gap", "Recommendation"]
            _add_header_row(ws_cm, headers)
            for e in cm_entries:
                cm_rows.append([
                    e.get("control_id", e.get("control_no", "")),
                    e.get("control_name", e.get("control_title", "")),
                    e.get("status", ""),
                    e.get("evidence", e.get("evidence_summary", "")),
                    e.get("gap", e.get("gap_description", "")),
                    e.get("recommendation", ""),
                ])
        _write_rows(ws_cm, cm_rows)
        _auto_width(ws_cm)

    # --- Sheet 4: Risk Register ---
    rr = fa.get("risk_register", {})
    risk_entries = []
    if isinstance(rr, dict):
        risk_entries = rr.get("generated_risk_entries", []) + rr.get("uploaded_risk_entries", [])
    elif isinstance(rr, list):
        risk_entries = rr

    valid_risks = []
    for r in risk_entries:
        rid = str(r.get("risk_id", r.get("id", "")))
        if rid.startswith("RSK-"): continue
        rThreat = str(r.get("threat", "")).lower()
        if not rThreat or rThreat in ("unspecified threat", "identified risk", "placeholder"):
            continue
        valid_risks.append(r)

    if valid_risks:
        ws_rr = wb.create_sheet("Risk Register")
        headers = ["Risk ID", "Risk Statement", "Asset", "Threat", "Likelihood", "Impact", "Risk Level", "Control", "Owner"]
        _add_header_row(ws_rr, headers)
        rows = []
        for i, r in enumerate(valid_risks):
            rid = str(r.get("risk_id", r.get("id", f"R{i+1}")))
            rStatement = r.get("risk_statement") or r.get("description") or r.get("title") or r.get("name") or r.get("risk_name") or "—"
            rAsset = r.get("asset") or r.get("asset_type") or "System"
            rThreat = r.get("threat") or r.get("vulnerability") or r.get("rule_id") or "Unspecified Threat"
            
            try:
                rawL = float(r.get("likelihood", 3))
                rawI = float(r.get("impact", 3))
            except:
                level = str(r.get("risk_level", r.get("level", "Medium"))).lower()
                if level == "low": rawL, rawI = 2, 2
                elif level == "high": rawL, rawI = 4, 3
                elif level in ("critical", "extreme"): rawL, rawI = 5, 5
                else: rawL, rawI = 3, 3
                
            lScore = max(1, min(5, int(round(rawL))))
            iScore = max(1, min(5, int(round(rawI))))
            
            score = lScore * iScore
            if score <= 5: risk_level = f"{score} Low"
            elif score <= 10: risk_level = f"{score} Medium"
            elif score <= 15: risk_level = f"{score} High"
            else: risk_level = f"{score} Extreme"
            
            ctrl = "—"
            if r.get("iso_controls") and isinstance(r["iso_controls"], list) and len(r["iso_controls"]) > 0:
                ctrl = ", ".join(r["iso_controls"])
            elif r.get("rule_id"): ctrl = r.get("rule_id")
            elif r.get("controls"): ctrl = r.get("controls")
            elif fa.get("framework"): ctrl = f"Mapped via {fa['framework']}"
            
            owner = r.get("owner")
            if not owner:
                t = str(rAsset).lower()
                if "server" in t or "infrastructure" in t or "host" in t: owner = "IT Team"
                elif "auth" in t or "identity" in t or "access" in t or "iam" in t: owner = "IT Security"
                elif "db" in t or "database" in t or "storage" in t: owner = "DBA"
                elif "network" in t or "firewall" in t or "router" in t: owner = "DevOps"
                elif "vendor" in t or "third-party" in t or "supplier" in t: owner = "IT Security"
                else: owner = "IT Team"
                
            rows.append([rid, rStatement, rAsset, rThreat, lScore, iScore, risk_level, ctrl, owner])
        _write_rows(ws_rr, rows)
        _auto_width(ws_rr)

    # --- Sheet 5: Treatment Plan ---
    tp_entries = fa.get("risk_treatment_plan")
    if tp_entries and isinstance(tp_entries, list):
        ws_tp = wb.create_sheet("Treatment Plan")
        headers = ["Risk ID", "Treatment", "Due Date"]
        _add_header_row(ws_tp, headers)
        rows = []
        for e in tp_entries:
            dd = e.get("due_date", "")
            if dd and "-" in dd and not "/" in dd:
                p = dd.split("-")
                if len(p) == 3: dd = f"{p[2]}/{p[1]}/{p[0]}"
            rows.append([
                e.get("risk_id", ""),
                e.get("treatment", e.get("action", "")),
                dd
            ])
        _write_rows(ws_tp, rows)
        _auto_width(ws_tp)

    # --- Sheet 6: Training Matrix & Tracker ---
    tm_backend = fa.get("training_matrix_generated", {})
    if isinstance(tm_backend, dict) and tm_backend.get("role_based_matrix"):
        ws_tm = wb.create_sheet("Training Matrix")
        headers = ["Role / Group", "Training Content", "Frequency", "Risk / Incident Driver"]
        _add_header_row(ws_tm, headers)
        rows = []
        for r in tm_backend["role_based_matrix"]:
            rows.append([
                r.get("role", ""),
                r.get("content", ""),
                r.get("frequency", ""),
                r.get("driver", "")
            ])
        _write_rows(ws_tm, rows)
        _auto_width(ws_tm)
        
        tracker = tm_backend.get("employee_tracker", tm_backend.get("employee_training_tracker", []))
        if tracker:
            ws_trk = wb.create_sheet("Employee Training Tracker")
            headers = ["Employee Name", "Role", "Assigned Training", "Status", "Last Training Date", "Next Due Date"]
            _add_header_row(ws_trk, headers)
            rows = []
            for r in tracker:
                def fmt_date(d):
                    if not d or d == "Not Available": return "—"
                    p = str(d).split("-")
                    if len(p) == 3 and len(p[0]) == 4: return f"{p[2]}/{p[1]}/{p[0]}"
                    return str(d)
                
                st = r.get("status", r.get("training_status", "Pending"))
                if st.lower() == "completed (on time)": st = "Completed (On Time)"
                elif st.lower() == "overdue": st = "Overdue"
                
                rows.append([
                    r.get("employee", r.get("employee_name", r.get("name", "Unknown"))),
                    r.get("role", "Employee"),
                    r.get("assigned_training", r.get("required_modules", "Security Awareness")),
                    st,
                    fmt_date(r.get("last_training_date")),
                    fmt_date(r.get("next_due_date"))
                ])
            _write_rows(ws_trk, rows)
            _auto_width(ws_trk)

    # --- Sheet 7: Governance Calendar ---
    gc = fa.get("governance_calendar_generated")
    if not gc:
        gc = fa.get("governance_calendar")
    
    if gc and isinstance(gc, list) and len(gc) > 0:
        ws_gc = wb.create_sheet("Governance Calendar")
        first = gc[0]
        if "governance_activity" in first or "month" in first:
            headers = ["Month", "Governance Activity"]
            _add_header_row(ws_gc, headers)
            rows = []
            for i, e in enumerate(gc):
                rows.append([
                    e.get("month", f"Month {i+1}"),
                    e.get("governance_activity", e.get("activity", "—"))
                ])
        else:
            headers = ["Activity", "Cadence", "Responsible", "Status", "Due Date", "Notes"]
            _add_header_row(ws_gc, headers)
            rows = []
            for e in gc:
                rows.append([
                    e.get("activity", e.get("task", e.get("event", ""))),
                    e.get("cadence", e.get("frequency", "")),
                    e.get("responsible", e.get("owner", "")),
                    e.get("status", ""),
                    e.get("due_date", e.get("next_due", "")),
                    e.get("notes", e.get("remarks", "")),
                ])
        _write_rows(ws_gc, rows)
        _auto_width(ws_gc)

    # --- Sheet 8: Vendor Checklist ---
    vc = fa.get("vendor_checklist")
    vc_entries = []
    if isinstance(vc, dict):
        vc_entries = vc.get("entries", vc.get("vendors", []))
    elif isinstance(vc, list):
        vc_entries = vc

    if vc_entries:
        ws_vc = wb.create_sheet("Vendor Checklist")
        agreement_label = "Agreement Signed (BAA)" if is_hipaa else "Agreement Signed (DPA)"
        headers = ["Vendor / Service", "Certifications / Compliance", agreement_label, "Encryption (At Rest / Transit)", "Security SLA (Breach / Uptime)", "Monitoring Frequency", "Risk Level"]
        _add_header_row(ws_vc, headers)
        rows = []
        for e in vc_entries:
            agreement  = e.get("agreement", e.get("agreement_signed", e.get("baa_status", e.get("baa_signed", "—"))))
            encryption = e.get("encryption", e.get("encryption_status", e.get("encryption_at_rest_transit", "—")))
            sla        = e.get("sla", e.get("security_sla", e.get("breach_uptime_sla", "—")))
            monitoring = e.get("monitoring", e.get("monitoring_frequency", "—"))
            riskLevel  = e.get("risk_level", e.get("riskLevel", "Unknown"))
            certs      = e.get("certifications", e.get("compliance_status", e.get("certifications_compliance", "—")))

            service = e.get("service_provided", "")
            vendor_val = e.get("vendor_name", e.get("vendor", e.get("name", "")))
            if service:
                vendor_val = f"{vendor_val} ({service})"
            
            rows.append([
                vendor_val,
                certs,
                agreement,
                encryption,
                sla,
                monitoring,
                riskLevel
            ])
        _write_rows(ws_vc, rows)
        _auto_width(ws_vc)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"assessment_export_{company}_{date_str}.xlsx"
    return buf, filename


# ---------------------------------------------------------------------------
# PDF export  (simple text-based PDF using only stdlib + openpyxl already available)
# ---------------------------------------------------------------------------

class _SimplePDF:
    """
    Minimal PDF generator using raw PDF operators.
    No external dependency (no reportlab / fpdf required).
    """

    def __init__(self):
        self._pages: list[list[str]] = []
        self._current_page: list[str] = []
        self._y = 750  # current y position (top of content area)
        self._page_bottom = 60
        self._line_height = 14
        self._font_size = 10

    def _new_page(self):
        if self._current_page:
            self._pages.append(self._current_page)
        self._current_page = []
        self._y = 750

    def _ensure_space(self, lines_needed: int = 1):
        if self._y - (lines_needed * self._line_height) < self._page_bottom:
            self._new_page()

    def add_title(self, text: str):
        self._ensure_space(3)
        self._current_page.append(f"BT /F1 16 Tf 50 {self._y} Td ({self._esc(text)}) Tj ET")
        self._y -= 28

    def add_heading(self, text: str):
        self._ensure_space(2)
        self._y -= 8  # spacing before heading
        self._current_page.append(f"BT /F1 12 Tf 50 {self._y} Td ({self._esc(text)}) Tj ET")
        self._y -= 20

    def add_line(self, text: str, indent: int = 50):
        # Split long lines
        max_chars = 95
        while len(text) > max_chars:
            cut = text[:max_chars].rfind(" ")
            if cut <= 0:
                cut = max_chars
            self._ensure_space()
            self._current_page.append(
                f"BT /F2 {self._font_size} Tf {indent} {self._y} Td ({self._esc(text[:cut])}) Tj ET"
            )
            self._y -= self._line_height
            text = text[cut:].lstrip()
        if text:
            self._ensure_space()
            self._current_page.append(
                f"BT /F2 {self._font_size} Tf {indent} {self._y} Td ({self._esc(text)}) Tj ET"
            )
            self._y -= self._line_height

    def add_blank(self):
        self._y -= 6

    @staticmethod
    def _esc(text: str) -> str:
        return (
            str(text)
            .replace("\\", "\\\\")
            .replace("(", "\\(")
            .replace(")", "\\)")
            .replace("\r", "")
            .replace("\n", " ")
        )

    def render(self) -> bytes:
        if self._current_page:
            self._pages.append(self._current_page)
            self._current_page = []

        objects: list[bytes] = []
        offsets: list[int] = []
        output = b""

        def add_obj(content: str) -> int:
            nonlocal output
            idx = len(objects) + 1
            offsets.append(len(output))
            obj_bytes = f"{idx} 0 obj\n{content}\nendobj\n".encode("latin-1", errors="replace")
            objects.append(obj_bytes)
            output += obj_bytes
            return idx

        header = b"%PDF-1.4\n"
        output += header

        # Font objects
        f1_id = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
        f2_id = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

        # Resource dict string
        resources = f"<< /Font << /F1 {f1_id} 0 R /F2 {f2_id} 0 R >> >>"

        page_ids = []
        for page_cmds in self._pages:
            stream_content = "\n".join(page_cmds)
            stream_id = add_obj(
                f"<< /Length {len(stream_content)} >>\nstream\n{stream_content}\nendstream"
            )
            page_id = add_obj(
                f"<< /Type /Page /Parent PAGES_REF"
                f" /MediaBox [0 0 595 842]"
                f" /Contents {stream_id} 0 R"
                f" /Resources {resources} >>"
            )
            page_ids.append(page_id)

        kids = " ".join(f"{pid} 0 R" for pid in page_ids)
        pages_id = add_obj(
            f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>"
        )

        # Patch parent references
        new_output = output.replace(b"PAGES_REF", f"{pages_id} 0 R".encode())
        output = new_output

        catalog_id = add_obj(f"<< /Type /Catalog /Pages {pages_id} 0 R >>")

        # xref
        xref_offset = len(output)
        xref = f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
        for off in offsets:
            xref += f"{off:010d} 00000 n \n"
        output += xref.encode("latin-1")
        output += f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\n".encode()
        output += f"startxref\n{xref_offset}\n%%EOF\n".encode()

        return output


def build_pdf_report(fa: dict) -> tuple[io.BytesIO, str]:
    """
    Build a PDF report from a framework_assessment dict.

    Returns (BytesIO buffer, suggested filename).
    """
    pdf = _SimplePDF()
    company = fa.get("assessment_name") or fa.get("_company_name") or "Assessment"
    framework = fa.get("framework") or fa.get("_top_framework") or ""
    date_str = _fmt_date(fa.get("created_at") or fa.get("_top_created_at", ""))
    score = fa.get("compliance_score", 0)

    # --- Title Page ---
    pdf.add_title(f"{framework} Assessment Report")
    pdf.add_blank()
    pdf.add_line(f"Company / Assessment: {company}")
    pdf.add_line(f"Framework: {framework}")
    pdf.add_line(f"Date: {date_str}")
    pdf.add_line(f"Compliance Score: {score}%")
    pdf.add_line(f"Total Controls: {fa.get('total_controls', 0)}")
    pdf.add_line(f"Compliant: {fa.get('compliant_controls', 0)}")
    pdf.add_line(f"Partial: {fa.get('partial_controls', 0)}")
    pdf.add_line(f"Missing: {fa.get('missing_controls', 0)}")
    pdf.add_blank()

    # --- Key Findings ---
    pdf.add_heading("Key Findings")
    insights = fa.get("insights", [])
    if insights:
        for ins in insights[:10]:
            pdf.add_line(f"- {ins}", indent=60)
    else:
        pdf.add_line("No key findings available.")
    pdf.add_blank()

    # --- High Risks Summary ---
    pdf.add_heading("High Risks Summary")
    top_risks = fa.get("top_missing_high_risk", [])
    if top_risks:
        for r in top_risks:
            pdf.add_line(
                f"- [{r.get('rule_id', '')}] {r.get('name', '')} ({r.get('section_name', '')})",
                indent=60,
            )
    else:
        pdf.add_line("No high-risk items identified.")
    pdf.add_blank()

    # --- Risk Register summary ---
    rr = fa.get("risk_register", {})
    risk_entries = []
    if isinstance(rr, dict):
        risk_entries = rr.get("generated_risk_entries", []) + rr.get("uploaded_risk_entries", [])
    elif isinstance(rr, list):
        risk_entries = rr
    if risk_entries:
        pdf.add_heading("Risk Register Overview")
        pdf.add_line(f"Total risks identified: {len(risk_entries)}")
        high_count = 0
        for r in risk_entries:
            try:
                if float(r.get("likelihood", 0)) * float(r.get("impact", 0)) >= 12:
                    high_count += 1
            except (ValueError, TypeError):
                pass
        pdf.add_line(f"High / Critical risks: {high_count}")
        pdf.add_blank()

    # --- Treatment Plan Summary ---
    tp = fa.get("risk_treatment_plan") or fa.get("treatment_plan")
    tp_entries = []
    if isinstance(tp, list):
        tp_entries = tp
    elif isinstance(tp, dict):
        tp_entries = tp.get("entries", tp.get("actions", []))
    if tp_entries:
        pdf.add_heading("Treatment Plan Summary")
        pdf.add_line(f"Total treatment actions: {len(tp_entries)}")
        for e in tp_entries[:5]:
            pdf.add_line(
                f"- [{e.get('risk_id', '')}] {(e.get('treatment', e.get('action', '')))[:120]}",
                indent=60,
            )
        if len(tp_entries) > 5:
            pdf.add_line(f"  ... and {len(tp_entries) - 5} more actions.")
        pdf.add_blank()

    # --- Governance Calendar summary ---
    gc = fa.get("governance_calendar") or fa.get("governance_calendar_generated")
    gc_entries = []
    if isinstance(gc, dict):
        gc_entries = gc.get("entries", gc.get("events", gc.get("items", [])))
    elif isinstance(gc, list):
        gc_entries = gc
    if gc_entries:
        pdf.add_heading("Governance Calendar")
        pdf.add_line(f"Total scheduled activities: {len(gc_entries)}")
        for e in gc_entries[:5]:
            pdf.add_line(
                f"- {e.get('activity', e.get('task', ''))} | {e.get('frequency', '')} | Owner: {e.get('owner', e.get('responsible', ''))}",
                indent=60,
            )
        if len(gc_entries) > 5:
            pdf.add_line(f"  ... and {len(gc_entries) - 5} more items.")
        pdf.add_blank()

    # --- Training Matrix summary ---
    tm = fa.get("training_matrix") or fa.get("training_matrix_generated")
    tm_entries = []
    if isinstance(tm, dict):
        tm_entries = tm.get("entries", tm.get("employees", tm.get("records", [])))
    elif isinstance(tm, list):
        tm_entries = tm
    if tm_entries:
        pdf.add_heading("Training Matrix")
        pdf.add_line(f"Total training records: {len(tm_entries)}")
        pdf.add_blank()

    # --- Vendor Checklist summary ---
    vc = fa.get("vendor_checklist")
    vc_entries = []
    if isinstance(vc, dict):
        vc_entries = vc.get("entries", vc.get("vendors", []))
    elif isinstance(vc, list):
        vc_entries = vc
    if vc_entries:
        pdf.add_heading("Vendor Checklist")
        pdf.add_line(f"Total vendors assessed: {len(vc_entries)}")
        for e in vc_entries[:5]:
            vendor_name = e.get("vendor_name", e.get("vendor", e.get("name", "")))
            status = e.get("status", e.get("compliance_status", ""))
            pdf.add_line(f"- {vendor_name}: {status}", indent=60)
        if len(vc_entries) > 5:
            pdf.add_line(f"  ... and {len(vc_entries) - 5} more vendors.")
        pdf.add_blank()

    # --- Footer ---
    pdf.add_blank()
    pdf.add_line(f"Generated by SmartISMS / Aegis.One on {date_str}")

    buf = io.BytesIO(pdf.render())
    buf.seek(0)
    safe_company = _safe_filename(company)
    filename = f"assessment_report_{safe_company}_{date_str}.pdf"
    return buf, filename
