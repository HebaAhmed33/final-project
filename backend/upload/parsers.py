"""
File parsers for uploaded content.
Each parser returns a structured dict or raises HTTPException on malformed input.
"""

import io
import json
import logging

from fastapi import HTTPException

from upload.header_normalizer import normalize_headers
from upload.sheet_classifier import classify_sheet, normalize_columns_for_type

logger = logging.getLogger("aegis.upload.parsers")

try:
    import yaml
except ImportError:
    yaml = None  # type: ignore[assignment]

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# Excel parser — Assessment mode (LEGACY — kept for backward compat)
# ---------------------------------------------------------------------------

def parse_excel(contents: bytes, filename: str) -> dict:
    """
    Parse an Excel workbook — SMART mode.

    Instead of requiring exact columns and rejecting the file, this parser:
      1. Loads the workbook
      2. Tries the active sheet first
      3. Normalizes headers via the alias map
      4. If no useful columns found on the active sheet, scans ALL sheets
      5. Accepts any row that has at least one non-empty mapped field
      6. Never rejects a valid Excel file — always returns what it can extract

    Returns
    -------
    dict  with keys: rows, headers, total_rows, imported_rows, skipped_rows,
                     errors, warnings, detected_type
    """
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    # Try active sheet first, then all sheets
    best_sheet = None
    best_headers = []
    best_useful_count = 0

    _USEFUL_COLUMNS = {"control_id", "control_name", "status", "owner",
                       "policy_name", "asset_name", "risk_name", "vendor_name"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if len(raw_rows) < 2:
            continue

        headers = normalize_headers(raw_rows[0])
        useful_count = len(_USEFUL_COLUMNS & set(headers))

        if useful_count > best_useful_count:
            best_useful_count = useful_count
            best_sheet = sheet_name
            best_headers = headers

    if best_sheet is None:
        # No sheet had any useful columns — still don't reject,
        # fall back to the active sheet and use whatever headers exist
        ws = wb.active
        if ws is None:
            wb.close()
            return {
                "rows": [],
                "headers": [],
                "total_rows": 0,
                "imported_rows": 0,
                "skipped_rows": 0,
                "errors": [],
                "warnings": ["Workbook has no active sheet."],
                "detected_type": "unknown",
            }
        raw_rows = list(ws.iter_rows(values_only=True))
        if len(raw_rows) < 2:
            wb.close()
            return {
                "rows": [],
                "headers": [],
                "total_rows": 0,
                "imported_rows": 0,
                "skipped_rows": 0,
                "errors": [],
                "warnings": ["No data rows found in the workbook."],
                "detected_type": "unknown",
            }
        best_headers = normalize_headers(raw_rows[0])
        best_sheet = ws.title
    else:
        ws = wb[best_sheet]
        raw_rows = list(ws.iter_rows(values_only=True))

    warnings = []
    if best_useful_count == 0:
        warnings.append(
            f"No standard control columns detected. Processing all data as-is. "
            f"Headers found: {', '.join(h for h in best_headers if h)}"
        )

    # Parse rows — accept any row that has at least one non-empty value
    imported: list[dict] = []
    skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(raw_rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            skipped += 1
            continue

        entry: dict = {}
        for col_idx, header in enumerate(best_headers):
            if not header:
                continue
            val = row[col_idx] if col_idx < len(row) else None
            entry[header] = str(val).strip() if val is not None else ""

        # Accept row if it has any non-empty value in a mapped field
        has_any_data = any(bool(v) for v in entry.values())
        if not has_any_data:
            skipped += 1
            continue

        imported.append(entry)

    wb.close()

    logger.info(
        f"parse_excel: sheet='{best_sheet}', "
        f"imported={len(imported)}, skipped={skipped}, "
        f"headers={[h for h in best_headers if h]}"
    )

    return {
        "rows": imported,
        "headers": [h for h in best_headers if h],
        "total_rows": len(raw_rows) - 1,
        "imported_rows": len(imported),
        "skipped_rows": skipped,
        "errors": errors,
        "warnings": warnings,
        "detected_type": "controls" if best_useful_count > 0 else "inferred",
    }


# ---------------------------------------------------------------------------
# Configuration file parsers
# ---------------------------------------------------------------------------

def parse_json_config(contents: bytes, filename: str) -> dict:
    """Parse a JSON configuration file."""
    try:
        data = json.loads(contents.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise HTTPException(status_code=400, detail=f"Malformed JSON: {exc}")

    if not isinstance(data, dict):
        raise HTTPException(
            status_code=400,
            detail="JSON config must be a top-level object (not an array or scalar).",
        )

    return data


def parse_yaml_config(contents: bytes, filename: str) -> dict:
    """Parse a YAML / YML configuration file."""
    if yaml is None:
        raise HTTPException(status_code=500, detail="pyyaml is not installed.")

    try:
        data = yaml.safe_load(contents.decode("utf-8"))
    except (yaml.YAMLError, UnicodeDecodeError) as exc:
        raise HTTPException(status_code=400, detail=f"Malformed YAML: {exc}")

    if not isinstance(data, dict):
        raise HTTPException(
            status_code=400,
            detail="YAML config must be a top-level mapping (not a list or scalar).",
        )

    return data


def parse_env_config(contents: bytes, filename: str) -> dict:
    """Parse a .env file into key-value pairs."""
    try:
        text = contents.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Cannot decode .env file: {exc}")

    data: dict[str, str] = {}
    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue  # silently skip malformed lines
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip("\"'")
        if key:
            data[key] = value

    if not data:
        raise HTTPException(status_code=400, detail=".env file contains no valid key=value pairs.")

    return data


def parse_config_file(contents: bytes, filename: str) -> tuple[dict, str]:
    """
    Dispatch to the correct parser based on extension.

    Returns (parsed_dict, detected_type_label).
    """
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""

    if ext == ".json":
        return parse_json_config(contents, filename), "json"
    elif ext in {".yaml", ".yml"}:
        return parse_yaml_config(contents, filename), "yaml"
    elif ext == ".env":
        return parse_env_config(contents, filename), "env"
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported config file type: {ext}")


# ---------------------------------------------------------------------------
# Multi-sheet Excel parser — Assessment mode (PRIMARY)
# ---------------------------------------------------------------------------

def parse_all_sheets(contents: bytes, filename: str) -> list[dict]:
    """
    Parse every sheet in an Excel workbook into a list of sheet dicts.

    Smart ingestion engine behavior:
      - Parses ALL sheets, not just one
      - Classifies each sheet by name, headers, AND data heuristics
      - Normalizes columns per detected sheet type
      - Never rejects a valid file for missing columns
      - Provides detailed detection metadata per sheet

    Each element contains:
        name               : sheet name as it appears in the workbook
        type               : classified sheet type (assets / controls / vendors / …)
        row_count          : number of non-empty data rows
        headers            : raw header strings
        normalized_headers : canonical column names after alias mapping
        rows               : list of row dicts keyed by canonical column names
        classification     : how the sheet was classified (by_name / by_headers / by_data / unknown)
        warnings           : any warnings about the sheet

    Raises HTTPException on unreadable workbooks.
    """
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    result: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))

        if not raw_rows:
            result.append({
                "name": sheet_name,
                "type": "unknown",
                "row_count": 0,
                "headers": [],
                "normalized_headers": [],
                "rows": [],
                "classification": "empty",
                "warnings": [],
            })
            continue

        raw_headers = raw_rows[0]
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        # Detect if the first row is actually a title/banner row (single merged cell)
        non_empty_headers = [h for h in headers if h]
        if len(non_empty_headers) <= 1 and len(raw_rows) > 2:
            # The first row might be a title row — try using row 2 as headers
            logger.info(
                f"Sheet '{sheet_name}': first row looks like a title row "
                f"(only {len(non_empty_headers)} non-empty cells). "
                f"Trying row 2 as headers."
            )
            headers = [str(h).strip() if h is not None else "" for h in raw_rows[1]]
            non_empty_headers = [h for h in headers if h]
            if len(non_empty_headers) >= 2:
                raw_rows = raw_rows[1:]  # shift: row 2 is now the header
            else:
                # Revert to original
                headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        # Prepare sample rows for data heuristic classification
        sample_rows = raw_rows[1:11] if len(raw_rows) > 1 else []

        # Classify the sheet using name + headers + data heuristic
        sheet_type = classify_sheet(sheet_name, headers, sample_rows)

        # Track how classification was determined
        from upload.sheet_classifier import classify_sheet_by_name, classify_sheet_by_headers
        if classify_sheet_by_name(sheet_name):
            classification_method = "by_name"
        elif classify_sheet_by_headers(headers):
            classification_method = "by_headers"
        elif sheet_type != "unknown":
            classification_method = "by_data"
        else:
            classification_method = "unknown"

        # Normalize column names for the detected type
        normalized_headers = normalize_columns_for_type(headers, sheet_type)

        # Parse data rows
        rows: list[dict] = []
        sheet_warnings: list[str] = []

        for row in raw_rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # skip fully empty rows
            entry: dict = {}
            for col_idx, canonical in enumerate(normalized_headers):
                if not canonical:
                    continue
                val = row[col_idx] if col_idx < len(row) else None
                entry[canonical] = str(val).strip() if val is not None else ""
            # Accept any row that has at least one non-empty field
            if any(bool(v) for v in entry.values()):
                rows.append(entry)

        # Generate warnings for partially mapped sheets
        if sheet_type == "unknown" and len(rows) > 0:
            sheet_warnings.append(
                f"Sheet '{sheet_name}' could not be auto-classified. "
                f"{len(rows)} rows were preserved as raw data."
            )

        logger.info(
            f"Sheet '{sheet_name}': type={sheet_type}, "
            f"classification={classification_method}, "
            f"rows={len(rows)}, "
            f"headers={[h for h in headers if h]}, "
            f"normalized={[h for h in normalized_headers if h]}"
        )

        result.append({
            "name": sheet_name,
            "type": sheet_type,
            "row_count": len(rows),
            "headers": [h for h in headers if h],
            "normalized_headers": normalized_headers,
            "rows": rows,
            "classification": classification_method,
            "warnings": sheet_warnings,
        })

    wb.close()

    # Log summary
    classified = [s for s in result if s["type"] != "unknown"]
    unclassified = [s for s in result if s["type"] == "unknown" and s["row_count"] > 0]
    logger.info(
        f"Workbook '{filename}': {len(result)} sheets total, "
        f"{len(classified)} classified, {len(unclassified)} unclassified with data"
    )

    return result


# Backward-compat alias
parse_excel_multisheet = parse_all_sheets
